from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

import xl_2_yaml
from residual.config import load_app_config


CORE_SHEETS = [
    "Catalog",
    "Task",
    "Dataset",
    "Runtime",
    "Training",
    "CV",
    "Scheduler",
    "Residual",
    "Jobs",
    "SearchSpace",
]


def _repo_with_data(tmp_path: Path) -> Path:
    repo_root = tmp_path
    (repo_root / "data").mkdir(parents=True, exist_ok=True)
    (repo_root / "data" / "df.csv").write_text(
        "dt,Com_CrudeOil,Com_Gasoline\n2024-01-01,1.0,2.0\n2024-01-08,1.1,2.1\n2024-01-15,1.2,2.2\n2024-01-22,1.3,2.3\n",
        encoding="utf-8",
    )
    return repo_root


def _row_index(sheet, catalog_id: str) -> int:
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == catalog_id:
            return row_idx
    raise AssertionError(f"catalog_id {catalog_id} not found in {sheet.title}")


def _set_row(sheet, catalog_id: str, values: dict[str, object]) -> None:
    headers = [cell.value for cell in sheet[1]]
    header_map = {str(name): idx + 1 for idx, name in enumerate(headers)}
    row_idx = sheet.max_row + 1
    sheet.cell(row=row_idx, column=1, value=catalog_id)
    for key, value in values.items():
        sheet.cell(row=row_idx, column=header_map[key], value=value)


def _build_valid_workbook(workbook_path: Path, *, family: str = "feature_set", stem: str = "case1") -> Path:
    xl_2_yaml.create_template_workbook(workbook_path)
    workbook = load_workbook(workbook_path)

    _set_row(
        workbook["Catalog"],
        "cfg1",
        {
            "family": family,
            "config_stem": stem,
            "file_target": f"{stem}.yaml",
            "adapter_set": family,
            "enabled": True,
        },
    )
    _set_row(workbook["Task"], "cfg1", {"name": f"{stem}_task"})
    _set_row(
        workbook["Dataset"],
        "cfg1",
        {
            "path": "data/df.csv",
            "target_col": "Com_CrudeOil",
            "dt_col": "dt",
            "hist_exog_cols": "Com_Gasoline",
            "futr_exog_cols": "",
            "static_exog_cols": "",
        },
    )
    _set_row(workbook["Runtime"], "cfg1", {"random_seed": 1})
    _set_row(
        workbook["Training"],
        "cfg1",
        {
            "input_size": 8,
            "season_length": 2,
            "batch_size": 4,
            "valid_batch_size": 4,
            "windows_batch_size": 16,
            "inference_windows_batch_size": 16,
            "learning_rate": 0.001,
            "max_steps": 5,
            "val_size": 1,
            "val_check_steps": 1,
            "early_stop_patience_steps": 1,
        },
    )
    _set_row(
        workbook["CV"],
        "cfg1",
        {"horizon": 1, "step_size": 1, "n_windows": 2, "gap": 0, "overlap_eval_policy": "by_cutoff_mean"},
    )
    _set_row(
        workbook["Scheduler"],
        "cfg1",
        {
            "gpu_ids": "0\n1",
            "max_concurrent_jobs": 2,
            "worker_devices": 1,
            "parallelize_single_job_tuning": True,
        },
    )
    _set_row(workbook["Residual"], "cfg1", {"enabled": False})
    _set_row(workbook["Jobs"], "cfg1", {"job_index": 1, "model": "Naive", "params": "{}"})

    workbook.save(workbook_path)
    return workbook_path


def test_create_template_workbook_contains_required_sheets(tmp_path: Path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    xl_2_yaml.create_template_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    for sheet_name in CORE_SHEETS:
        assert sheet_name in workbook.sheetnames
    assert "Adapter.feature_set" in workbook.sheetnames
    assert "Adapter.bomb_trans" in workbook.sheetnames


def test_generate_from_workbook_routes_yaml_to_family_directory(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "case.xlsx", family="univar", stem="baseline-case")

    result = xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)

    generated = repo_root / "yaml" / "univar" / "baseline-case.yaml"
    assert result.generated_paths == (generated,)
    assert generated.exists()
    loaded = load_app_config(repo_root, config_path=generated)
    assert loaded.config.task.name == "baseline-case_task"
    assert loaded.config.dataset.target_col == "Com_CrudeOil"


def test_generate_rejects_unknown_family(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "unknown-family.xlsx", family="feature_set", stem="family-case")
    workbook = load_workbook(workbook_path)
    row_idx = _row_index(workbook["Catalog"], "cfg1")
    headers = [cell.value for cell in workbook["Catalog"][1]]
    workbook["Catalog"].cell(row=row_idx, column=headers.index("family") + 1, value="mystery_family")
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="Unsupported family"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=False)


def test_generate_from_workbook_rejects_colliding_targets(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "collision.xlsx", family="feature_set", stem="dup")
    workbook = load_workbook(workbook_path)

    _set_row(
        workbook["Catalog"],
        "cfg2",
        {
            "family": "feature_set",
            "config_stem": "dup",
            "file_target": "dup.yaml",
            "adapter_set": "feature_set",
            "enabled": True,
        },
    )
    _set_row(workbook["Task"], "cfg2", {"name": "dup_two"})
    _set_row(workbook["Dataset"], "cfg2", {"path": "data/df.csv", "target_col": "Com_CrudeOil", "dt_col": "dt"})
    _set_row(workbook["Runtime"], "cfg2", {"random_seed": 1})
    _set_row(workbook["Training"], "cfg2", {"input_size": 8})
    _set_row(workbook["CV"], "cfg2", {"horizon": 1, "step_size": 1, "n_windows": 2, "gap": 0, "overlap_eval_policy": "by_cutoff_mean"})
    _set_row(workbook["Scheduler"], "cfg2", {"gpu_ids": "0\n1", "max_concurrent_jobs": 2, "worker_devices": 1, "parallelize_single_job_tuning": True})
    _set_row(workbook["Residual"], "cfg2", {"enabled": False})
    _set_row(workbook["Jobs"], "cfg2", {"job_index": 1, "model": "Naive", "params": "{}"})
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="same target path"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=False)

    assert not (repo_root / "yaml").exists()


def test_generate_from_workbook_rolls_back_on_validation_error(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "invalid.xlsx", family="feature_set", stem="bad-one")
    workbook = load_workbook(workbook_path)

    _set_row(
        workbook["Catalog"],
        "cfg2",
        {
            "family": "feature_set",
            "config_stem": "bad-two",
            "file_target": "bad-two.yaml",
            "adapter_set": "feature_set",
            "enabled": True,
        },
    )
    _set_row(workbook["Task"], "cfg2", {"name": "bad_two"})
    _set_row(workbook["Dataset"], "cfg2", {"path": "data/df.csv", "target_col": "Com_CrudeOil", "dt_col": "dt"})
    _set_row(workbook["Runtime"], "cfg2", {"random_seed": 1, "transformations_target": "bad"})
    _set_row(workbook["Training"], "cfg2", {"input_size": 8})
    _set_row(workbook["CV"], "cfg2", {"horizon": 1, "step_size": 1, "n_windows": 2, "gap": 0, "overlap_eval_policy": "by_cutoff_mean"})
    _set_row(workbook["Scheduler"], "cfg2", {"gpu_ids": "0\n1", "max_concurrent_jobs": 2, "worker_devices": 1, "parallelize_single_job_tuning": True})
    _set_row(workbook["Residual"], "cfg2", {"enabled": False})
    _set_row(workbook["Jobs"], "cfg2", {"job_index": 1, "model": "Naive", "params": "{}"})
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="Validation failed"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)

    assert not list((repo_root / "yaml").rglob("*.yaml"))


@pytest.mark.parametrize(
    ("sheet_name", "column_name"),
    [
        ("Runtime", "transformations"),
        ("CV", "final_holdout"),
        ("Residual", "train_source"),
    ],
)
def test_generate_rejects_removed_legacy_workbook_columns(
    tmp_path: Path, sheet_name: str, column_name: str
) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / f"{sheet_name}.xlsx", family="feature_set", stem="legacy-case")
    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name]
    extra_col = sheet.max_column + 1
    sheet.cell(row=1, column=extra_col, value=column_name)
    sheet.cell(row=2, column=extra_col, value="legacy")
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="unsupported column"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=False)


def test_reverse_round_trip_preserves_semantics_and_explicit_defaults(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    config_path = repo_root / "yaml" / "feature_set" / "orig.yaml"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    original_payload = {
        "task": {"name": "orig_task"},
        "dataset": {
            "path": "data/df.csv",
            "target_col": "Com_CrudeOil",
            "dt_col": "dt",
            "hist_exog_cols": ["Com_Gasoline"],
            "futr_exog_cols": [],
            "static_exog_cols": [],
        },
        "runtime": {"random_seed": 1},
        "training": {
            "input_size": 8,
            "season_length": 2,
            "batch_size": 4,
            "valid_batch_size": 4,
            "windows_batch_size": 16,
            "inference_windows_batch_size": 16,
            "learning_rate": 0.001,
            "max_steps": 5,
            "val_size": 1,
            "val_check_steps": 1,
            "early_stop_patience_steps": 1,
            "loss": "mse",
        },
        "cv": {"horizon": 1, "step_size": 1, "n_windows": 2, "gap": 0, "overlap_eval_policy": "by_cutoff_mean"},
        "scheduler": {"gpu_ids": [0, 1], "max_concurrent_jobs": 2, "worker_devices": 1, "parallelize_single_job_tuning": True},
        "residual": {"enabled": False},
        "jobs": [{"model": "Naive", "params": {}}],
    }
    config_path.write_text(yaml.safe_dump(original_payload, sort_keys=False), encoding="utf-8")
    original_loaded = load_app_config(repo_root, config_path=config_path).config.to_dict()

    workbook_path = tmp_path / "reverse.xlsx"
    xl_2_yaml.reverse_yaml_to_workbook([config_path], output_path=workbook_path, repo_root=repo_root)
    workbook = load_workbook(workbook_path)
    training_sheet = workbook["Training"]
    row_idx = _row_index(training_sheet, "cfg_001")
    headers = [cell.value for cell in training_sheet[1]]
    loss_col = headers.index("loss") + 1
    protocol_col = headers.index("train_protocol") + 1
    assert training_sheet.cell(row=row_idx, column=loss_col).value == "mse"
    assert training_sheet.cell(row=row_idx, column=protocol_col).value is None

    xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)
    regenerated_text = config_path.read_text(encoding="utf-8")
    regenerated_loaded = load_app_config(repo_root, config_path=config_path).config.to_dict()
    assert regenerated_loaded == original_loaded
    assert "loss: mse" in regenerated_text
    assert "train_protocol:" not in regenerated_text


def test_generate_rejects_non_allowlisted_adapter_override(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "adapter.xlsx", family="feature_set", stem="adapter-case")
    workbook = load_workbook(workbook_path)
    _set_row(
        workbook["Adapter.feature_set"],
        "cfg1",
        {
            "enabled": True,
            "field_path": "jobs.0.model",
            "value_json": '"LSTM"',
        },
    )
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="override allowlist"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=False)


def test_adapter_set_binds_only_matching_adapter_sheet(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    workbook_path = _build_valid_workbook(tmp_path / "adapter-binding.xlsx", family="feature_set", stem="binding-case")
    workbook = load_workbook(workbook_path)
    _set_row(
        workbook["Adapter.bomb"],
        "cfg1",
        {
            "enabled": True,
            "field_path": "task.name",
            "value_json": '"wrong_name"',
        },
    )
    workbook.save(workbook_path)

    result = xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)
    generated = result.generated_paths[0]
    loaded = load_app_config(repo_root, config_path=generated)
    assert loaded.config.task.name == "binding-case_task"


def test_auto_mode_requires_matching_searchspace_rows(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    (repo_root / "search_space.yaml").write_text(
        yaml.safe_dump(
            {
                "models": {"LSTM": ["encoder_hidden_size"]},
                "training": [],
                "residual": {"xgboost": []},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = _build_valid_workbook(tmp_path / "auto.xlsx", family="feature_set", stem="auto-case")
    workbook = load_workbook(workbook_path)
    jobs_sheet = workbook["Jobs"]
    row_idx = _row_index(jobs_sheet, "cfg1")
    headers = [cell.value for cell in jobs_sheet[1]]
    jobs_sheet.cell(row=row_idx, column=headers.index("model") + 1, value="LSTM")
    jobs_sheet.cell(row=row_idx, column=headers.index("params") + 1, value="{}")
    _set_row(
        workbook["SearchSpace"],
        "cfg1",
        {"section": "models", "model": "LSTM", "selectors": "encoder_hidden_size"},
    )
    workbook.save(workbook_path)

    result = xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)
    generated = result.generated_paths[0]
    loaded = load_app_config(repo_root, config_path=generated)
    assert loaded.config.jobs[0].validated_mode == "learned_auto"


def test_auto_mode_searchspace_mismatch_fails(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    (repo_root / "search_space.yaml").write_text(
        yaml.safe_dump(
            {
                "models": {"LSTM": ["encoder_hidden_size"]},
                "training": [],
                "residual": {"xgboost": []},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = _build_valid_workbook(tmp_path / "auto-mismatch.xlsx", family="feature_set", stem="auto-mismatch")
    workbook = load_workbook(workbook_path)
    jobs_sheet = workbook["Jobs"]
    row_idx = _row_index(jobs_sheet, "cfg1")
    headers = [cell.value for cell in jobs_sheet[1]]
    jobs_sheet.cell(row=row_idx, column=headers.index("model") + 1, value="LSTM")
    jobs_sheet.cell(row=row_idx, column=headers.index("params") + 1, value="{}")
    workbook.save(workbook_path)

    with pytest.raises(xl_2_yaml.XL2YAMLError, match="SearchSpace sheet mismatch"):
        xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)


def test_reverse_auto_mode_round_trip_populates_searchspace_sheet(tmp_path: Path) -> None:
    repo_root = _repo_with_data(tmp_path)
    (repo_root / "search_space.yaml").write_text(
        yaml.safe_dump(
            {
                "models": {"LSTM": ["encoder_hidden_size"]},
                "training": [],
                "residual": {"xgboost": []},
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    config_path = repo_root / "yaml" / "feature_set" / "auto-roundtrip.yaml"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        yaml.safe_dump(
            {
                "task": {"name": "auto_roundtrip"},
                "dataset": {
                    "path": "data/df.csv",
                    "target_col": "Com_CrudeOil",
                    "dt_col": "dt",
                    "hist_exog_cols": ["Com_Gasoline"],
                    "futr_exog_cols": [],
                    "static_exog_cols": [],
                },
                "runtime": {"random_seed": 1},
                "training": {
                    "input_size": 8,
                    "season_length": 2,
                    "batch_size": 4,
                    "valid_batch_size": 4,
                    "windows_batch_size": 16,
                    "inference_windows_batch_size": 16,
                    "learning_rate": 0.001,
                    "max_steps": 5,
                    "val_size": 1,
                    "val_check_steps": 1,
                    "early_stop_patience_steps": 1,
                },
                "cv": {"horizon": 1, "step_size": 1, "n_windows": 2, "gap": 0, "overlap_eval_policy": "by_cutoff_mean"},
                "scheduler": {"gpu_ids": [0, 1], "max_concurrent_jobs": 2, "worker_devices": 1, "parallelize_single_job_tuning": True},
                "residual": {"enabled": False},
                "jobs": [{"model": "LSTM", "params": {}}],
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    workbook_path = tmp_path / "auto-roundtrip.xlsx"
    xl_2_yaml.reverse_yaml_to_workbook([config_path], output_path=workbook_path, repo_root=repo_root)
    workbook = load_workbook(workbook_path)
    search_sheet = workbook["SearchSpace"]
    row_idx = _row_index(search_sheet, "cfg_001")
    headers = [cell.value for cell in search_sheet[1]]
    section_col = headers.index("section") + 1
    model_col = headers.index("model") + 1
    selectors_col = headers.index("selectors") + 1
    assert search_sheet.cell(row=row_idx, column=section_col).value == "models"
    assert search_sheet.cell(row=row_idx, column=model_col).value == "LSTM"
    assert search_sheet.cell(row=row_idx, column=selectors_col).value == "encoder_hidden_size"

    xl_2_yaml.generate_from_workbook(workbook_path, repo_root=repo_root, validate=True)
    loaded = load_app_config(repo_root, config_path=config_path)
    assert loaded.config.jobs[0].validated_mode == "learned_auto"
    assert tuple(loaded.config.jobs[0].selected_search_params) == ("encoder_hidden_size",)


def test_parse_args_treats_plain_workbook_path_as_generate() -> None:
    args = xl_2_yaml.parse_args(["template.xlsx"])
    assert args.command == "generate"
    assert args.workbook == Path("template.xlsx")
