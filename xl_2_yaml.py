from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from residual.config import load_app_config
from residual.runtime import (
    _build_resolved_artifacts,
    _selected_jobs,
    _validate_adapters,
    _validate_jobs,
)

REPO_ROOT = Path(__file__).resolve().parent
YAML_ROOT = REPO_ROOT / "yaml"
DOCUMENTED_FAMILIES = (
    "feature_set",
    "feature_set_HPT",
    "feature_set_HPT_c3",
    "feature_set_HPT_n100",
    "feature_set_residual",
    "bomb",
    "bomb_trans",
    "univar",
    "blackswan",
    "jaeho_feature_set",
)

TRUE_VALUES = {True, 1, "1", "true", "yes", "y", "on"}
FALSE_VALUES = {False, 0, "0", "false", "no", "n", "off"}


class XL2YAMLError(RuntimeError):
    pass


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    kind: str


@dataclass(frozen=True)
class CatalogEntry:
    catalog_id: str
    family: str
    config_stem: str
    file_target: str
    adapter_set: str | None
    enabled: bool


@dataclass(frozen=True)
class ValidationReport:
    path: Path
    ok: bool
    detail: str


@dataclass(frozen=True)
class RuntimeValidationResult:
    report: ValidationReport
    loaded: Any | None
    selected_jobs: tuple[Any, ...]


@dataclass(frozen=True)
class GenerationResult:
    generated_paths: tuple[Path, ...]
    validation_reports: tuple[ValidationReport, ...]


SECTION_SPECS: dict[str, tuple[ColumnSpec, ...]] = {
    "Task": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("name", "str"),
    ),
    "Dataset": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("path", "str"),
        ColumnSpec("target_col", "str"),
        ColumnSpec("dt_col", "str"),
        ColumnSpec("freq", "str"),
        ColumnSpec("hist_exog_cols", "list[str]"),
        ColumnSpec("futr_exog_cols", "list[str]"),
        ColumnSpec("static_exog_cols", "list[str]"),
    ),
    "Runtime": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("random_seed", "int"),
        ColumnSpec("opt_n_trial", "int"),
        ColumnSpec("transformations_target", "str"),
        ColumnSpec("transformations_exog", "str"),
    ),
    "Training": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("train_protocol", "str"),
        ColumnSpec("input_size", "int"),
        ColumnSpec("season_length", "int"),
        ColumnSpec("batch_size", "int"),
        ColumnSpec("valid_batch_size", "int"),
        ColumnSpec("windows_batch_size", "int"),
        ColumnSpec("inference_windows_batch_size", "int"),
        ColumnSpec("learning_rate", "float"),
        ColumnSpec("scaler_type", "str"),
        ColumnSpec("model_step_size", "int"),
        ColumnSpec("max_steps", "int"),
        ColumnSpec("val_size", "int"),
        ColumnSpec("val_check_steps", "int"),
        ColumnSpec("early_stop_patience_steps", "int"),
        ColumnSpec("num_lr_decays", "int"),
        ColumnSpec("loss", "str"),
        ColumnSpec("loss_params", "json"),
        ColumnSpec("accelerator", "str"),
        ColumnSpec("devices", "int"),
        ColumnSpec("strategy", "str"),
        ColumnSpec("precision", "scalar"),
        ColumnSpec("dataloader_kwargs", "json"),
    ),
    "CV": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("horizon", "int"),
        ColumnSpec("step_size", "int"),
        ColumnSpec("n_windows", "int"),
        ColumnSpec("gap", "int"),
        ColumnSpec("max_train_size", "int"),
        ColumnSpec("overlap_eval_policy", "str"),
    ),
    "Scheduler": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("gpu_ids", "list[int]"),
        ColumnSpec("max_concurrent_jobs", "int"),
        ColumnSpec("worker_devices", "int"),
        ColumnSpec("parallelize_single_job_tuning", "bool"),
    ),
    "Residual": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("enabled", "bool"),
        ColumnSpec("model", "str"),
        ColumnSpec("target", "str"),
        ColumnSpec("cpu_threads", "int"),
        ColumnSpec("params", "json"),
        ColumnSpec("features_include_base_prediction", "bool"),
        ColumnSpec("features_include_horizon_step", "bool"),
        ColumnSpec("features_include_date_features", "bool"),
        ColumnSpec("features_lag_enabled", "bool"),
        ColumnSpec("features_lag_sources", "list[str]"),
        ColumnSpec("features_lag_steps", "list[int]"),
        ColumnSpec("features_lag_transforms", "list[str]"),
        ColumnSpec("features_exog_hist", "list[str]"),
        ColumnSpec("features_exog_futr", "list[str]"),
        ColumnSpec("features_exog_static", "list[str]"),
    ),
    "Jobs": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("job_index", "int"),
        ColumnSpec("model", "str"),
        ColumnSpec("params", "json"),
    ),
    "SearchSpace": (
        ColumnSpec("catalog_id", "str"),
        ColumnSpec("section", "str"),
        ColumnSpec("model", "str"),
        ColumnSpec("selectors", "list[str]"),
    ),
}

CATALOG_SPECS = (
    ColumnSpec("catalog_id", "str"),
    ColumnSpec("family", "str"),
    ColumnSpec("file_target", "str"),
    ColumnSpec("config_stem", "str"),
    ColumnSpec("adapter_set", "str"),
    ColumnSpec("enabled", "bool"),
)

ADAPTER_SPECS = (
    ColumnSpec("catalog_id", "str"),
    ColumnSpec("enabled", "bool"),
    ColumnSpec("field_path", "str"),
    ColumnSpec("value_json", "scalar"),
)

SECTION_CANONICAL_FIELDS: dict[str, tuple[str, ...]] = {
    "Task": ("name",),
    "Dataset": ("path", "target_col", "dt_col", "freq", "hist_exog_cols", "futr_exog_cols", "static_exog_cols"),
    "Runtime": ("random_seed", "opt_n_trial", "transformations_target", "transformations_exog"),
    "Training": (
        "train_protocol", "input_size", "season_length", "batch_size", "valid_batch_size",
        "windows_batch_size", "inference_windows_batch_size", "learning_rate", "scaler_type",
        "model_step_size", "max_steps", "val_size", "val_check_steps",
        "early_stop_patience_steps", "num_lr_decays", "loss", "loss_params",
        "accelerator", "devices", "strategy", "precision", "dataloader_kwargs",
    ),
    "CV": ("horizon", "step_size", "n_windows", "gap", "max_train_size", "overlap_eval_policy"),
    "Scheduler": ("gpu_ids", "max_concurrent_jobs", "worker_devices", "parallelize_single_job_tuning"),
    "Residual": (
        "enabled", "model", "target", "cpu_threads", "params",
        "features_include_base_prediction", "features_include_horizon_step",
        "features_include_date_features", "features_lag_enabled", "features_lag_sources",
        "features_lag_steps", "features_lag_transforms", "features_exog_hist",
        "features_exog_futr", "features_exog_static",
    ),
}

ADAPTER_OVERRIDE_ALLOWLIST = {
    f"task.{field}" for field in SECTION_CANONICAL_FIELDS["Task"]
} | {
    f"dataset.{field}" for field in SECTION_CANONICAL_FIELDS["Dataset"]
} | {
    f"runtime.{field}" for field in SECTION_CANONICAL_FIELDS["Runtime"]
} | {
    f"training.{field}" for field in SECTION_CANONICAL_FIELDS["Training"]
} | {
    f"cv.{field}" for field in SECTION_CANONICAL_FIELDS["CV"]
} | {
    f"scheduler.{field}" for field in SECTION_CANONICAL_FIELDS["Scheduler"]
} | {
    f"residual.{field}" for field in SECTION_CANONICAL_FIELDS["Residual"]
}


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_bool(value: Any) -> bool | None:
    if _is_blank(value):
        return None
    normalized = value.strip().lower() if isinstance(value, str) else value
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise XL2YAMLError(f"Cannot parse boolean value: {value!r}")


def _parse_json_or_scalar(value: Any) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    text = str(value).strip()
    if not text:
        return None
    if text[0] in "[{\"" or text in {"true", "false", "null"}:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text
    return text


def _split_text_list(text: str) -> list[str]:
    if "\n" in text:
        parts = [part.strip() for part in text.splitlines()]
    else:
        parts = [part.strip() for part in text.split(",")]
    return [part for part in parts if part]


def _parse_list(value: Any, item_kind: str) -> list[Any] | None:
    if _is_blank(value):
        return None
    if isinstance(value, list):
        items = value
    else:
        parsed = _parse_json_or_scalar(value)
        if isinstance(parsed, list):
            items = parsed
        elif isinstance(parsed, str):
            items = _split_text_list(parsed)
        else:
            items = [parsed]
    out: list[Any] = []
    for item in items:
        if item_kind == "str":
            out.append(str(item))
        elif item_kind == "int":
            out.append(int(item))
        else:
            raise XL2YAMLError(f"Unsupported list item kind: {item_kind}")
    return out


def _parse_scalar(value: Any, kind: str) -> Any:
    if _is_blank(value):
        return None
    if kind == "str":
        return str(value).strip()
    if kind == "int":
        return int(value)
    if kind == "float":
        return float(value)
    if kind == "bool":
        return _parse_bool(value)
    if kind == "json":
        parsed = _parse_json_or_scalar(value)
        if parsed is None:
            return None
        if isinstance(parsed, str):
            try:
                return json.loads(parsed)
            except json.JSONDecodeError as exc:
                raise XL2YAMLError(f"Expected JSON object/list, got: {value!r}") from exc
        return parsed
    if kind == "scalar":
        return _parse_json_or_scalar(value)
    if kind.startswith("list["):
        item_kind = kind[5:-1]
        return _parse_list(value, item_kind)
    raise XL2YAMLError(f"Unsupported column kind: {kind}")


def _serialize_cell(value: Any, kind: str) -> Any:
    if value is None:
        return None
    if kind.startswith("list["):
        return "\n".join(str(item) for item in value)
    if kind == "json":
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if kind == "bool":
        return bool(value)
    return value


def _normalize_family(family: str) -> str:
    normalized = family.strip()
    if not normalized:
        raise XL2YAMLError("Catalog.family is required")
    if "/" in normalized or normalized.startswith("."):
        raise XL2YAMLError(f"Unsafe family name: {family!r}")
    return normalized


def _safe_relative_target(value: str) -> Path:
    candidate = Path(value)
    if candidate.is_absolute() or any(part == ".." for part in candidate.parts):
        raise XL2YAMLError(f"Unsafe file_target: {value!r}")
    return candidate


def _yaml_target_for_catalog(repo_root: Path, entry: CatalogEntry) -> Path:
    family_root = repo_root / "yaml" / entry.family
    if entry.file_target:
        relative = _safe_relative_target(entry.file_target)
        if relative.suffix.lower() not in {".yaml", ".yml"}:
            relative = relative.with_suffix(".yaml")
    else:
        relative = Path(f"{entry.config_stem}.yaml")
    return family_root / relative


def _add_validation(sheet, family_col: str | None, bool_cols: list[str]) -> None:
    if family_col is not None:
        family_dv = DataValidation(type="list", formula1='"' + ",".join(DOCUMENTED_FAMILIES) + '"', allow_blank=False)
        sheet.add_data_validation(family_dv)
        family_dv.add(f"{family_col}2:{family_col}500")
    for column in bool_cols:
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        sheet.add_data_validation(dv)
        dv.add(f"{column}2:{column}500")


def _format_sheet(sheet, headers: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        width = max(12, min(36, len(header) + 2))
        sheet.column_dimensions[get_column_letter(idx)].width = width


def create_template_workbook(output_path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    catalog = workbook.create_sheet("Catalog")
    catalog_headers = [spec.key for spec in CATALOG_SPECS]
    _format_sheet(catalog, catalog_headers)
    _add_validation(catalog, family_col="B", bool_cols=["F"])

    for sheet_name, specs in SECTION_SPECS.items():
        sheet = workbook.create_sheet(sheet_name)
        _format_sheet(sheet, [spec.key for spec in specs])
        bool_cols = [get_column_letter(idx + 1) for idx, spec in enumerate(specs) if spec.kind == "bool"]
        if bool_cols:
            _add_validation(sheet, family_col=None, bool_cols=bool_cols)

    for family in DOCUMENTED_FAMILIES:
        sheet = workbook.create_sheet(f"Adapter.{family}")
        headers = [spec.key for spec in ADAPTER_SPECS]
        _format_sheet(sheet, headers)
        _add_validation(sheet, family_col=None, bool_cols=["B"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _load_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        raise XL2YAMLError(f"Workbook is missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(not _is_blank(value) for value in row):
            continue
        rows.append({str(headers[idx]): value for idx, value in enumerate(row) if headers[idx] is not None})
    return rows


def _read_workbook(
    path: Path,
) -> tuple[
    dict[str, CatalogEntry],
    dict[str, dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    dict[str, dict[str, list[dict[str, Any]]]],
]:
    workbook = load_workbook(path)
    required_sheets = {"Catalog", *SECTION_SPECS.keys()}
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise XL2YAMLError(f"Workbook is missing required sheets: {', '.join(sorted(missing))}")

    catalog_entries: dict[str, CatalogEntry] = {}
    catalog_rows = _load_sheet_rows(path, "Catalog")
    if not catalog_rows:
        raise XL2YAMLError("Catalog sheet must contain at least one row")
    for raw in catalog_rows:
        catalog_id = str(raw.get("catalog_id", "")).strip()
        if not catalog_id:
            raise XL2YAMLError("Catalog.catalog_id is required")
        if catalog_id in catalog_entries:
            raise XL2YAMLError(f"Duplicate catalog_id: {catalog_id}")
        family = _normalize_family(str(raw.get("family", "")))
        config_stem = str(raw.get("config_stem", "")).strip() or str(raw.get("file_target", "")).strip()
        if not config_stem:
            raise XL2YAMLError(f"Catalog row {catalog_id} must define config_stem or file_target")
        file_target = str(raw.get("file_target", "")).strip()
        adapter_set = str(raw.get("adapter_set", "")).strip() or None
        enabled = _parse_bool(raw.get("enabled"))
        catalog_entries[catalog_id] = CatalogEntry(
            catalog_id=catalog_id,
            family=family,
            config_stem=config_stem,
            file_target=file_target,
            adapter_set=adapter_set,
            enabled=True if enabled is None else enabled,
        )

    single_rows: dict[str, dict[str, Any]] = {sheet: {} for sheet in SECTION_SPECS if sheet not in {"Jobs", "SearchSpace"}}
    multi_rows: dict[str, list[dict[str, Any]]] = {"Jobs": [], "SearchSpace": []}
    for sheet_name, specs in SECTION_SPECS.items():
        rows = _load_sheet_rows(path, sheet_name)
        for raw in rows:
            catalog_id = str(raw.get("catalog_id", "")).strip()
            if catalog_id not in catalog_entries:
                raise XL2YAMLError(f"Sheet {sheet_name} references unknown catalog_id: {catalog_id}")
            parsed = {
                spec.key: _parse_scalar(raw.get(spec.key), spec.kind)
                for spec in specs
            }
            if sheet_name in multi_rows:
                multi_rows[sheet_name].append(parsed)
            else:
                if catalog_id in single_rows[sheet_name]:
                    raise XL2YAMLError(f"Sheet {sheet_name} has duplicate row for catalog_id={catalog_id}")
                single_rows[sheet_name][catalog_id] = parsed

    adapter_rows: dict[str, dict[str, list[dict[str, Any]]]] = {
        catalog_id: {} for catalog_id in catalog_entries
    }
    for family in DOCUMENTED_FAMILIES:
        sheet_name = f"Adapter.{family}"
        if sheet_name not in workbook.sheetnames:
            continue
        for raw in _load_sheet_rows(path, sheet_name):
            catalog_id = str(raw.get("catalog_id", "")).strip()
            if catalog_id not in catalog_entries:
                raise XL2YAMLError(f"Sheet {sheet_name} references unknown catalog_id: {catalog_id}")
            enabled = _parse_bool(raw.get("enabled"))
            field_path = str(raw.get("field_path", "")).strip()
            value = raw.get("value_json")
            adapter_rows[catalog_id].setdefault(family, []).append(
                {
                    "catalog_id": catalog_id,
                    "enabled": True if enabled is None else enabled,
                    "field_path": field_path,
                    "value": _parse_json_or_scalar(value),
                }
            )

    for catalog_id, entry in catalog_entries.items():
        if not entry.enabled:
            continue
        for sheet_name in single_rows:
            if catalog_id not in single_rows[sheet_name]:
                raise XL2YAMLError(f"Catalog row {catalog_id} is missing required binding in sheet {sheet_name}")
        job_rows = [row for row in multi_rows["Jobs"] if row["catalog_id"] == catalog_id]
        if not job_rows:
            raise XL2YAMLError(f"Catalog row {catalog_id} must define at least one Jobs row")

    return catalog_entries, single_rows, multi_rows, adapter_rows


def _strip_nones(payload: Any) -> Any:
    if isinstance(payload, dict):
        out = {key: _strip_nones(value) for key, value in payload.items()}
        return {key: value for key, value in out.items() if value not in (None, {}, [])}
    if isinstance(payload, list):
        return [_strip_nones(item) for item in payload]
    return payload


def _set_nested(payload: dict[str, Any], dotted_path: str, value: Any) -> None:
    parts = dotted_path.split(".")
    cursor = payload
    for part in parts[:-1]:
        cursor = cursor.setdefault(part, {})
    cursor[parts[-1]] = value


def _build_section_payload(section: str, row: dict[str, Any]) -> dict[str, Any]:
    data = {key: value for key, value in row.items() if key != "catalog_id" and value is not None}
    if section == "Residual":
        feature_payload: dict[str, Any] = {}
        lag_payload: dict[str, Any] = {}
        exog_payload: dict[str, Any] = {}
        for key in list(data):
            if key.startswith("features_lag_"):
                lag_payload[key.removeprefix("features_lag_")] = data.pop(key)
            elif key.startswith("features_exog_"):
                exog_payload[key.removeprefix("features_exog_")] = data.pop(key)
            elif key.startswith("features_"):
                feature_payload[key.removeprefix("features_")] = data.pop(key)
        if lag_payload:
            feature_payload["lag_features"] = lag_payload
        if exog_payload:
            feature_payload["exog_sources"] = {
                key: value for key, value in exog_payload.items() if value is not None
            }
        if feature_payload:
            data["features"] = feature_payload
    return _strip_nones(data)


def _apply_adapters(payload: dict[str, Any], adapter_rows: list[dict[str, Any]]) -> None:
    for row in adapter_rows:
        if not row["enabled"]:
            continue
        field_path = row["field_path"]
        if not field_path:
            continue
        if field_path not in ADAPTER_OVERRIDE_ALLOWLIST:
            raise XL2YAMLError(
                f"Adapter field_path {field_path!r} is not on the override allowlist"
            )
        value = row["value"]
        if value is None:
            continue
        _set_nested(payload, field_path, value)


def _build_payload_for_catalog(
    catalog_id: str,
    entry: CatalogEntry,
    single_rows: dict[str, dict[str, Any]],
    multi_rows: dict[str, list[dict[str, Any]]],
    adapter_rows: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    section_map = {
        "Task": "task",
        "Dataset": "dataset",
        "Runtime": "runtime",
        "Training": "training",
        "CV": "cv",
        "Scheduler": "scheduler",
        "Residual": "residual",
    }
    for sheet_name, section_name in section_map.items():
        payload_section = _build_section_payload(sheet_name, single_rows[sheet_name][catalog_id])
        if payload_section:
            payload[section_name] = payload_section
    jobs = [
        {
            "model": row["model"],
            "params": row["params"] or {},
        }
        for row in sorted(
            (row for row in multi_rows["Jobs"] if row["catalog_id"] == catalog_id),
            key=lambda row: row["job_index"] or 0,
        )
    ]
    payload["jobs"] = jobs
    adapter_key = entry.adapter_set or entry.family
    _apply_adapters(payload, adapter_rows.get(adapter_key, []))
    stripped = _strip_nones(payload)
    normalized_jobs = []
    for job in stripped.get("jobs", []):
        normalized_job = dict(job)
        normalized_job.setdefault("params", {})
        normalized_jobs.append(normalized_job)
    stripped["jobs"] = normalized_jobs
    return stripped


def _validate_no_collisions(entries: dict[str, CatalogEntry], repo_root: Path) -> dict[str, Path]:
    targets: dict[str, Path] = {}
    seen: dict[Path, str] = {}
    for catalog_id, entry in entries.items():
        if not entry.enabled:
            continue
        target = _yaml_target_for_catalog(repo_root, entry)
        if target in seen:
            raise XL2YAMLError(
                f"Catalog rows {seen[target]} and {catalog_id} resolve to the same target path: {target}"
            )
        seen[target] = catalog_id
        targets[catalog_id] = target
    return targets


def _runtime_validate_generated_config(
    repo_root: Path, config_path: Path
) -> RuntimeValidationResult:
    temp_root = Path(tempfile.mkdtemp(prefix="xl2yaml-validate-", dir=repo_root))
    try:
        loaded = load_app_config(repo_root, config_path=config_path)
        output_root = temp_root / "validation-output"
        paths = _build_resolved_artifacts(repo_root, loaded, output_root)
        selected_jobs = tuple(_selected_jobs(loaded, None))
        _validate_jobs(loaded, selected_jobs, paths["capability_path"])
        _validate_adapters(loaded, selected_jobs)
        return RuntimeValidationResult(
            report=ValidationReport(path=config_path, ok=True, detail="runtime validate-only ok"),
            loaded=loaded,
            selected_jobs=selected_jobs,
        )
    except Exception as exc:  # noqa: BLE001
        return RuntimeValidationResult(
            report=ValidationReport(path=config_path, ok=False, detail=str(exc)),
            loaded=None,
            selected_jobs=(),
        )
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def _expected_search_space_rows(loaded: Any, selected_jobs: tuple[Any, ...]) -> set[tuple[str, str, tuple[str, ...]]]:
    rows: set[tuple[str, str, tuple[str, ...]]] = set()
    training_selectors = tuple(loaded.config.training_search.selected_search_params)
    if training_selectors:
        rows.add(("training", "", training_selectors))
    for job in selected_jobs:
        selectors = tuple(job.selected_search_params)
        if selectors:
            rows.add(("models", job.model, selectors))
    residual_selectors = tuple(loaded.config.residual.selected_search_params)
    if residual_selectors:
        rows.add(("residual", loaded.config.residual.model, residual_selectors))
    return rows


def _declared_search_space_rows(
    catalog_id: str, multi_rows: dict[str, list[dict[str, Any]]]
) -> set[tuple[str, str, tuple[str, ...]]]:
    rows: set[tuple[str, str, tuple[str, ...]]] = set()
    for row in multi_rows["SearchSpace"]:
        if row["catalog_id"] != catalog_id:
            continue
        section = str(row.get("section") or "").strip().lower()
        model = str(row.get("model") or "").strip()
        selectors = tuple(row.get("selectors") or ())
        rows.add((section, model, selectors))
    return rows


def _validate_search_space_contract(
    catalog_id: str, multi_rows: dict[str, list[dict[str, Any]]], loaded: Any, selected_jobs: tuple[Any, ...]
) -> None:
    declared = _declared_search_space_rows(catalog_id, multi_rows)
    expected = _expected_search_space_rows(loaded, selected_jobs)
    if declared != expected:
        raise XL2YAMLError(
            f"SearchSpace sheet mismatch for {catalog_id}: declared={sorted(declared)!r} expected={sorted(expected)!r}"
        )


def _default_validation_runner(repo_root: Path, config_path: Path) -> ValidationReport:
    return _runtime_validate_generated_config(repo_root, config_path).report


def generate_from_workbook(
    workbook_path: Path,
    *,
    repo_root: Path = REPO_ROOT,
    catalog_ids: set[str] | None = None,
    validate: bool = True,
    validation_runner: Callable[[Path, Path], RuntimeValidationResult] | None = None,
) -> GenerationResult:
    repo_root = repo_root.resolve()
    catalog_entries, single_rows, multi_rows, adapter_rows = _read_workbook(workbook_path)
    if catalog_ids is not None:
        catalog_entries = {
            catalog_id: entry
            for catalog_id, entry in catalog_entries.items()
            if catalog_id in catalog_ids
        }
        if not catalog_entries:
            raise XL2YAMLError("No catalog rows matched the requested catalog_ids")
        single_rows = {
            sheet: {catalog_id: row for catalog_id, row in rows.items() if catalog_id in catalog_entries}
            for sheet, rows in single_rows.items()
        }
        multi_rows = {
            sheet: [row for row in rows if row["catalog_id"] in catalog_entries]
            for sheet, rows in multi_rows.items()
        }
        adapter_rows = {
            catalog_id: rows
            for catalog_id, rows in adapter_rows.items()
            if catalog_id in catalog_entries
        }
    targets = _validate_no_collisions(catalog_entries, repo_root)
    validation_runner = validation_runner or _runtime_validate_generated_config
    generated_paths: list[Path] = []
    validation_reports: list[ValidationReport] = []

    temp_root = Path(tempfile.mkdtemp(prefix="xl2yaml-", dir=repo_root))
    try:
        staged_paths: dict[str, Path] = {}
        for catalog_id, entry in catalog_entries.items():
            if not entry.enabled:
                continue
            payload = _build_payload_for_catalog(
                catalog_id, entry, single_rows, multi_rows, adapter_rows.get(catalog_id, {})
            )
            staged_path = temp_root / targets[catalog_id].relative_to(repo_root)
            staged_path.parent.mkdir(parents=True, exist_ok=True)
            staged_path.write_text(
                yaml.safe_dump(payload, sort_keys=False, allow_unicode=True),
                encoding="utf-8",
            )
            staged_paths[catalog_id] = staged_path
        if validate:
            for catalog_id, staged_path in staged_paths.items():
                validation = validation_runner(repo_root, staged_path)
                validation_reports.append(validation.report)
                if not validation.report.ok:
                    raise XL2YAMLError(
                        f"Validation failed for {catalog_id} ({targets[catalog_id]}): {validation.report.detail}"
                    )
                _validate_search_space_contract(
                    catalog_id, multi_rows, validation.loaded, validation.selected_jobs
                )
        for catalog_id, staged_path in staged_paths.items():
            target = targets[catalog_id]
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(staged_path), str(target))
            generated_paths.append(target)
        return GenerationResult(tuple(generated_paths), tuple(validation_reports))
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)


def _relative_family_path(repo_root: Path, config_path: Path) -> tuple[str, str, str]:
    try:
        relative = config_path.resolve().relative_to((repo_root / "yaml").resolve())
    except ValueError as exc:
        raise XL2YAMLError(
            f"YAML path {config_path} must live under {repo_root / 'yaml'} for reverse conversion"
        ) from exc
    if len(relative.parts) < 2:
        raise XL2YAMLError(f"YAML path {config_path} is missing a family directory")
    family = relative.parts[0]
    file_target = str(Path(*relative.parts[1:]))
    return family, Path(file_target).stem, file_target


def _row_map(specs: tuple[ColumnSpec, ...], values: dict[str, Any]) -> list[Any]:
    row: list[Any] = []
    for spec in specs:
        row.append(_serialize_cell(values.get(spec.key), spec.kind))
    return row


def _raw_section(payload: dict[str, Any], key: str) -> dict[str, Any]:
    section = payload.get(key)
    return dict(section) if isinstance(section, dict) else {}


def reverse_yaml_to_workbook(
    config_paths: list[Path],
    *,
    output_path: Path,
    repo_root: Path = REPO_ROOT,
) -> Path:
    create_template_workbook(output_path)
    workbook = load_workbook(output_path)
    sheets = {name: workbook[name] for name in workbook.sheetnames}

    for index, config_path in enumerate(config_paths, start=1):
        payload = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
        if not isinstance(payload, dict):
            raise XL2YAMLError(f"Config {config_path} must be a mapping")
        loaded = load_app_config(repo_root, config_path=config_path)
        family, config_stem, file_target = _relative_family_path(repo_root, config_path)
        catalog_id = f"cfg_{index:03d}"
        sheets["Catalog"].append([
            catalog_id,
            family,
            file_target,
            config_stem,
            family,
            True,
        ])

        task = _raw_section(payload, "task")
        sheets["Task"].append(_row_map(SECTION_SPECS["Task"], {
            "catalog_id": catalog_id,
            "name": task.get("name"),
        }))

        dataset = _raw_section(payload, "dataset")
        sheets["Dataset"].append(_row_map(SECTION_SPECS["Dataset"], {
            "catalog_id": catalog_id,
            "path": dataset.get("path"),
            "target_col": dataset.get("target_col"),
            "dt_col": dataset.get("dt_col"),
            "freq": dataset.get("freq"),
            "hist_exog_cols": dataset.get("hist_exog_cols"),
            "futr_exog_cols": dataset.get("futr_exog_cols"),
            "static_exog_cols": dataset.get("static_exog_cols"),
        }))

        runtime = _raw_section(payload, "runtime")
        sheets["Runtime"].append(_row_map(SECTION_SPECS["Runtime"], {
            "catalog_id": catalog_id,
            "random_seed": runtime.get("random_seed"),
            "opt_n_trial": runtime.get("opt_n_trial"),
            "transformations_target": runtime.get("transformations_target"),
            "transformations_exog": runtime.get("transformations_exog"),
        }))

        training = _raw_section(payload, "training")
        sheets["Training"].append(_row_map(SECTION_SPECS["Training"], {
            "catalog_id": catalog_id,
            **training,
        }))

        cv = _raw_section(payload, "cv")
        sheets["CV"].append(_row_map(SECTION_SPECS["CV"], {
            "catalog_id": catalog_id,
            **cv,
        }))

        scheduler = _raw_section(payload, "scheduler")
        sheets["Scheduler"].append(_row_map(SECTION_SPECS["Scheduler"], {
            "catalog_id": catalog_id,
            **scheduler,
        }))

        residual = _raw_section(payload, "residual")
        residual_features = dict(residual.get("features", {}) or {})
        lag_features = dict(residual_features.get("lag_features", {}) or {})
        exog_sources = dict(residual_features.get("exog_sources", {}) or {})
        sheets["Residual"].append(_row_map(SECTION_SPECS["Residual"], {
            "catalog_id": catalog_id,
            "enabled": residual.get("enabled"),
            "model": residual.get("model"),
            "target": residual.get("target"),
            "cpu_threads": residual.get("cpu_threads"),
            "params": residual.get("params"),
            "features_include_base_prediction": residual_features.get("include_base_prediction"),
            "features_include_horizon_step": residual_features.get("include_horizon_step"),
            "features_include_date_features": residual_features.get("include_date_features"),
            "features_lag_enabled": lag_features.get("enabled"),
            "features_lag_sources": lag_features.get("sources"),
            "features_lag_steps": lag_features.get("steps"),
            "features_lag_transforms": lag_features.get("transforms"),
            "features_exog_hist": exog_sources.get("hist"),
            "features_exog_futr": exog_sources.get("futr"),
            "features_exog_static": exog_sources.get("static"),
        }))

        for job_index, job in enumerate(payload.get("jobs", []) or [], start=1):
            if not isinstance(job, dict):
                continue
            sheets["Jobs"].append(_row_map(SECTION_SPECS["Jobs"], {
                "catalog_id": catalog_id,
                "job_index": job_index,
                "model": job.get("model"),
                "params": job.get("params"),
            }))

        for section, model, selectors in sorted(
            _expected_search_space_rows(loaded, tuple(loaded.config.jobs))
        ):
            sheets["SearchSpace"].append(
                _row_map(
                    SECTION_SPECS["SearchSpace"],
                    {
                        "catalog_id": catalog_id,
                        "section": section,
                        "model": model,
                        "selectors": list(selectors),
                    },
                )
            )

    workbook.save(output_path)
    return output_path


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel template <-> YAML workflow helper")
    subparsers = parser.add_subparsers(dest="command")

    template = subparsers.add_parser("template", help="Create a blank workbook template")
    template.add_argument("output", type=Path)

    generate = subparsers.add_parser("generate", help="Generate YAML files from a workbook")
    generate.add_argument("workbook", type=Path)
    generate.add_argument("--catalog-id", dest="catalog_ids", action="append", default=[])
    generate.add_argument("--repo-root", type=Path, default=REPO_ROOT)
    generate.add_argument("--no-validate", action="store_true")

    reverse = subparsers.add_parser("reverse", help="Reverse existing YAML files into a workbook")
    reverse.add_argument("configs", nargs="+", type=Path)
    reverse.add_argument("--output", required=True, type=Path)
    reverse.add_argument("--repo-root", type=Path, default=REPO_ROOT)

    return parser


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    argv = list(sys.argv[1:] if argv is None else argv)
    commands = {"template", "generate", "reverse"}
    if argv and argv[0] not in commands and not argv[0].startswith("-"):
        argv = ["generate", *argv]
    parser = _parser()
    args = parser.parse_args(argv)
    if args.command is None:
        parser.print_help()
        raise SystemExit(2)
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.command == "template":
        path = create_template_workbook(args.output)
        print(path)
        return 0
    if args.command == "generate":
        result = generate_from_workbook(
            args.workbook,
            repo_root=args.repo_root,
            catalog_ids=set(args.catalog_ids) or None,
            validate=not args.no_validate,
        )
        for path in result.generated_paths:
            print(path)
        return 0
    if args.command == "reverse":
        path = reverse_yaml_to_workbook(args.configs, output_path=args.output, repo_root=args.repo_root)
        print(path)
        return 0
    raise SystemExit(2)


if __name__ == "__main__":
    raise SystemExit(main())
